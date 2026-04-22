"""
生成 2000 行 × 300 列的 Excel 演示数据，适用于多因子相关性分析系统测试。
包含制造过程参数、环境参数、质量指标等多种因子，并内置已知相关性关系。
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

np.random.seed(42)

ROWS = 2000
COLS = 300


def generate_data():
    """生成 300 列数据，分为多个主题组，组内存在已知相关性。"""
    data = np.zeros((ROWS, COLS))
    col_idx = 0

    # ── 基础索引列 (1-5) ──
    data[:, col_idx] = np.arange(1, ROWS + 1, dtype=float)          # 样本编号
    col_idx += 1
    data[:, col_idx] = np.random.randint(1, 51, ROWS).astype(float) # 批次号
    col_idx += 1
    data[:, col_idx] = np.random.randint(1, 11, ROWS).astype(float) # 设备编号
    col_idx += 1
    data[:, col_idx] = np.random.randint(1, 5, ROWS).astype(float)  # 班次
    col_idx += 1
    data[:, col_idx] = np.random.uniform(20, 35, ROWS)              # 环境温度(基准)
    col_idx += 1  # col 5

    # ── 涂布工序参数 (6-45, 40列) ──
    # 基础参数
    slurry_viscosity = np.random.normal(5000, 500, ROWS)
    slurry_solid_content = np.random.normal(45, 2, ROWS)
    coat_speed = np.random.normal(50, 5, ROWS)
    coat_gap = np.random.normal(150, 10, ROWS)
    dry_temp = np.random.normal(110, 8, ROWS)
    dry_speed = np.random.normal(30, 3, ROWS)
    tension = np.random.normal(100, 10, ROWS)

    data[:, col_idx] = slurry_viscosity; col_idx += 1
    data[:, col_idx] = slurry_solid_content; col_idx += 1
    data[:, col_idx] = coat_speed; col_idx += 1
    data[:, col_idx] = coat_gap; col_idx += 1
    data[:, col_idx] = dry_temp; col_idx += 1
    data[:, col_idx] = dry_speed; col_idx += 1
    data[:, col_idx] = tension; col_idx += 1

    # 衍生参数（与基础参数有相关性）
    coat_thickness = 0.3 * coat_gap + 0.02 * slurry_solid_content + np.random.normal(0, 5, ROWS)
    coat_weight = coat_thickness * slurry_solid_content / 100 + np.random.normal(0, 0.5, ROWS)
    coat_density = coat_weight / coat_thickness + np.random.normal(0, 0.01, ROWS)
    surface_roughness = 0.1 * coat_speed + 0.05 * tension + np.random.normal(0, 0.5, ROWS)
    dry_rate = 0.5 * dry_temp + 0.8 * dry_speed + np.random.normal(0, 2, ROWS)
    residual_moisture = 50 - 0.3 * dry_temp - 0.5 * dry_speed + np.random.normal(0, 1, ROWS)
    porosity = 0.02 * coat_thickness + 0.01 * coat_weight + np.random.normal(0, 0.5, ROWS)
    adhesion = 10 - 0.001 * surface_roughness + 0.05 * residual_moisture + np.random.normal(0, 0.3, ROWS)

    data[:, col_idx] = coat_thickness; col_idx += 1
    data[:, col_idx] = coat_weight; col_idx += 1
    data[:, col_idx] = coat_density; col_idx += 1
    data[:, col_idx] = surface_roughness; col_idx += 1
    data[:, col_idx] = dry_rate; col_idx += 1
    data[:, col_idx] = residual_moisture; col_idx += 1
    data[:, col_idx] = porosity; col_idx += 1
    data[:, col_idx] = adhesion; col_idx += 1

    # 更多涂布参数（部分独立，部分弱相关）
    for i in range(25):
        if i % 3 == 0:
            data[:, col_idx] = slurry_viscosity * 0.01 * (i + 1) + np.random.normal(0, 2, ROWS)
        elif i % 3 == 1:
            data[:, col_idx] = coat_speed * 0.05 * (i + 1) + np.random.normal(0, 1.5, ROWS)
        else:
            data[:, col_idx] = np.random.normal(100 + i * 10, 15, ROWS)
        col_idx += 1  # 到 col 45

    # ── 辊压工序参数 (46-85, 40列) ──
    roll_pressure = np.random.normal(300, 30, ROWS)
    roll_speed = np.random.normal(20, 3, ROWS)
    roll_temp = np.random.normal(80, 5, ROWS)
    roll_gap = np.random.normal(100, 8, ROWS)
    nip_angle = np.random.normal(15, 1, ROWS)

    data[:, col_idx] = roll_pressure; col_idx += 1
    data[:, col_idx] = roll_speed; col_idx += 1
    data[:, col_idx] = roll_temp; col_idx += 1
    data[:, col_idx] = roll_gap; col_idx += 1
    data[:, col_idx] = nip_angle; col_idx += 1

    pressed_density = 2.0 + 0.003 * roll_pressure - 0.001 * roll_gap + np.random.normal(0, 0.1, ROWS)
    pressed_thickness = 0.5 * roll_gap + 0.002 * roll_speed + np.random.normal(0, 2, ROWS)
    elongation = 0.01 * roll_pressure + 0.05 * roll_speed + np.random.normal(0, 0.3, ROWS)
    smoothness = -0.05 * surface_roughness + 0.01 * roll_pressure + np.random.normal(0, 0.5, ROWS)
    hardness = 0.02 * roll_pressure + 0.01 * pressed_density + np.random.normal(0, 1, ROWS)
    conductivity = 0.5 * pressed_density + 0.001 * roll_temp + np.random.normal(0, 0.2, ROWS)

    data[:, col_idx] = pressed_density; col_idx += 1
    data[:, col_idx] = pressed_thickness; col_idx += 1
    data[:, col_idx] = elongation; col_idx += 1
    data[:, col_idx] = smoothness; col_idx += 1
    data[:, col_idx] = hardness; col_idx += 1
    data[:, col_idx] = conductivity; col_idx += 1

    for i in range(29):
        if i % 4 == 0:
            data[:, col_idx] = roll_pressure * 0.02 * (i + 1) + np.random.normal(0, 3, ROWS)
        elif i % 4 == 1:
            data[:, col_idx] = pressed_density * (i + 1) * 0.5 + np.random.normal(0, 2, ROWS)
        elif i % 4 == 2:
            data[:, col_idx] = roll_speed * 0.1 * (i + 1) + np.random.normal(0, 1, ROWS)
        else:
            data[:, col_idx] = np.random.normal(50 + i * 5, 8, ROWS)
        col_idx += 1  # 到 col 85

    # ── 装配工序参数 (86-125, 40列) ──
    weld_current = np.random.normal(200, 15, ROWS)
    weld_voltage = np.random.normal(12, 1, ROWS)
    weld_time = np.random.normal(50, 5, ROWS)
    weld_pressure_asm = np.random.normal(150, 12, ROWS)
    tab_width = np.random.normal(20, 0.5, ROWS)
    ultrasonic_power = np.random.normal(1000, 80, ROWS)

    data[:, col_idx] = weld_current; col_idx += 1
    data[:, col_idx] = weld_voltage; col_idx += 1
    data[:, col_idx] = weld_time; col_idx += 1
    data[:, col_idx] = weld_pressure_asm; col_idx += 1
    data[:, col_idx] = tab_width; col_idx += 1
    data[:, col_idx] = ultrasonic_power; col_idx += 1

    weld_resistance = 5 - 0.01 * weld_current - 0.1 * weld_time + np.random.normal(0, 0.5, ROWS)
    weld_strength = 0.05 * weld_current + 0.1 * weld_pressure_asm - 0.01 * weld_resistance + np.random.normal(0, 0.3, ROWS)
    seal_width = 0.1 * weld_time + 0.001 * weld_current + np.random.normal(0, 0.2, ROWS)
    seal_depth = 0.05 * weld_pressure_asm + 0.02 * weld_time + np.random.normal(0, 0.1, ROWS)
    joint_torque = 0.01 * ultrasonic_power + 0.05 * weld_pressure_asm + np.random.normal(0, 0.5, ROWS)

    data[:, col_idx] = weld_resistance; col_idx += 1
    data[:, col_idx] = weld_strength; col_idx += 1
    data[:, col_idx] = seal_width; col_idx += 1
    data[:, col_idx] = seal_depth; col_idx += 1
    data[:, col_idx] = joint_torque; col_idx += 1

    for i in range(29):
        if i % 3 == 0:
            data[:, col_idx] = weld_current * 0.05 * (i + 1) + np.random.normal(0, 2, ROWS)
        elif i % 3 == 1:
            data[:, col_idx] = ultrasonic_power * 0.002 * (i + 1) + np.random.normal(0, 1, ROWS)
        else:
            data[:, col_idx] = np.random.normal(200 + i * 20, 25, ROWS)
        col_idx += 1  # 到 col 125

    # ── 注液工序参数 (126-155, 30列) ──
    inj_volume = np.random.normal(50, 2, ROWS)
    inj_speed = np.random.normal(10, 1, ROWS)
    inj_pressure = np.random.normal(0.5, 0.05, ROWS)
    inj_temp = np.random.normal(25, 2, ROWS)
    vacuum_level = np.random.normal(-0.09, 0.005, ROWS)
    hold_time = np.random.normal(30, 3, ROWS)

    data[:, col_idx] = inj_volume; col_idx += 1
    data[:, col_idx] = inj_speed; col_idx += 1
    data[:, col_idx] = inj_pressure; col_idx += 1
    data[:, col_idx] = inj_temp; col_idx += 1
    data[:, col_idx] = vacuum_level; col_idx += 1
    data[:, col_idx] = hold_time; col_idx += 1

    wetting_time = 5 + 0.1 * inj_volume - 0.5 * inj_speed + np.random.normal(0, 1, ROWS)
    absorption_rate = 0.8 + 0.01 * inj_pressure + 0.02 * inj_temp + np.random.normal(0, 0.05, ROWS)
    retention = 0.5 * hold_time + 10 * vacuum_level + np.random.normal(0, 0.3, ROWS)

    data[:, col_idx] = wetting_time; col_idx += 1
    data[:, col_idx] = absorption_rate; col_idx += 1
    data[:, col_idx] = retention; col_idx += 1

    for i in range(21):
        data[:, col_idx] = np.random.normal(100 + i * 8, 12, ROWS)
        col_idx += 1  # 到 col 155

    # ── 化成工序参数 (156-185, 30列) ──
    form_current = np.random.normal(0.5, 0.05, ROWS)
    form_voltage = np.random.normal(3.7, 0.2, ROWS)
    form_temp = np.random.normal(45, 3, ROWS)
    form_time = np.random.normal(24, 2, ROWS)
    cut_voltage = np.random.normal(2.5, 0.1, ROWS)

    data[:, col_idx] = form_current; col_idx += 1
    data[:, col_idx] = form_voltage; col_idx += 1
    data[:, col_idx] = form_temp; col_idx += 1
    data[:, col_idx] = form_time; col_idx += 1
    data[:, col_idx] = cut_voltage; col_idx += 1

    sei_resistance = 0.5 * form_current + 0.1 * form_time + np.random.normal(0, 0.2, ROWS)
    capacity_initial = 100 + 20 * form_time - 5 * form_temp + 10 * inj_volume + np.random.normal(0, 3, ROWS)
    coulombic_eff = 98 - 0.5 * sei_resistance + 0.01 * form_time + np.random.normal(0, 0.5, ROWS)

    data[:, col_idx] = sei_resistance; col_idx += 1
    data[:, col_idx] = capacity_initial; col_idx += 1
    data[:, col_idx] = coulombic_eff; col_idx += 1

    for i in range(22):
        data[:, col_idx] = np.random.normal(50 + i * 5, 8, ROWS)
        col_idx += 1  # 到 col 185

    # ── 品质检测指标 (186-225, 40列) ──
    ocv = 3.7 + 0.05 * pressed_density + 0.01 * capacity_initial + np.random.normal(0, 0.05, ROWS)
    impedance = 50 - 0.5 * pressed_density + 2 * sei_resistance + np.random.normal(0, 3, ROWS)
    capacity_rated = 0.95 * capacity_initial + np.random.normal(0, 2, ROWS)
    self_discharge = 2 + 0.01 * impedance - 0.001 * ocv + np.random.normal(0, 0.3, ROWS)
    cycle_100 = 0.98 * capacity_rated + np.random.normal(0, 1, ROWS)
    cycle_500 = 0.92 * capacity_rated - 0.05 * impedance + np.random.normal(0, 2, ROWS)
    thickness_change = 0.1 * elongation + 0.05 * porosity + np.random.normal(0, 1, ROWS)
    weight = 45 + 0.1 * coat_weight + 0.02 * pressed_thickness + np.random.normal(0, 0.5, ROWS)

    data[:, col_idx] = ocv; col_idx += 1
    data[:, col_idx] = impedance; col_idx += 1
    data[:, col_idx] = capacity_rated; col_idx += 1
    data[:, col_idx] = self_discharge; col_idx += 1
    data[:, col_idx] = cycle_100; col_idx += 1
    data[:, col_idx] = cycle_500; col_idx += 1
    data[:, col_idx] = thickness_change; col_idx += 1
    data[:, col_idx] = weight; col_idx += 1

    for i in range(32):
        if i % 4 == 0:
            data[:, col_idx] = capacity_rated * (0.9 - 0.001 * i) + np.random.normal(0, 1.5, ROWS)
        elif i % 4 == 1:
            data[:, col_idx] = impedance * (1 + 0.02 * i) + np.random.normal(0, 2, ROWS)
        elif i % 4 == 2:
            data[:, col_idx] = ocv - 0.001 * i + np.random.normal(0, 0.02, ROWS)
        else:
            data[:, col_idx] = np.random.normal(100, 15, ROWS)
        col_idx += 1  # 到 col 225

    # ── 环境监测参数 (226-255, 30列) ──
    env_humidity = np.random.normal(50, 10, ROWS)
    env_temp = np.random.normal(25, 3, ROWS)
    env_dust = np.random.normal(100, 20, ROWS)
    env_pressure = np.random.normal(101.3, 0.5, ROWS)
    env_vibration = np.random.normal(0.5, 0.1, ROWS)

    data[:, col_idx] = env_humidity; col_idx += 1
    data[:, col_idx] = env_temp; col_idx += 1
    data[:, col_idx] = env_dust; col_idx += 1
    data[:, col_idx] = env_pressure; col_idx += 1
    data[:, col_idx] = env_vibration; col_idx += 1

    env_heat_index = 0.5 * env_temp + 0.01 * env_humidity + np.random.normal(0, 0.5, ROWS)
    env_dew_point = env_temp - 0.2 * (100 - env_humidity) + np.random.normal(0, 1, ROWS)

    data[:, col_idx] = env_heat_index; col_idx += 1
    data[:, col_idx] = env_dew_point; col_idx += 1

    for i in range(23):
        data[:, col_idx] = np.random.normal(30 + i * 3, 5, ROWS)
        col_idx += 1  # 到 col 255

    # ── 原材料参数 (256-300, 45列) ──
    raw_particle_size = np.random.normal(10, 2, ROWS)
    raw_bet_surface = np.random.normal(5, 0.5, ROWS)
    raw_purity = np.random.normal(99.9, 0.05, ROWS)
    raw_ph = np.random.normal(7, 0.3, ROWS)
    raw_moisture = np.random.normal(0.5, 0.1, ROWS)
    raw_conductivity_raw = np.random.normal(1000, 100, ROWS)

    data[:, col_idx] = raw_particle_size; col_idx += 1
    data[:, col_idx] = raw_bet_surface; col_idx += 1
    data[:, col_idx] = raw_purity; col_idx += 1
    data[:, col_idx] = raw_ph; col_idx += 1
    data[:, col_idx] = raw_moisture; col_idx += 1
    data[:, col_idx] = raw_conductivity_raw; col_idx += 1

    raw_dispersity = raw_particle_size * 0.1 + raw_bet_surface * 0.2 + np.random.normal(0, 0.3, ROWS)
    raw_flowability = 20 - 0.5 * raw_particle_size - 0.1 * raw_moisture + np.random.normal(0, 1, ROWS)
    raw_tap_density = 1.5 + 0.05 * raw_particle_size - 0.02 * raw_bet_surface + np.random.normal(0, 0.1, ROWS)

    data[:, col_idx] = raw_dispersity; col_idx += 1
    data[:, col_idx] = raw_flowability; col_idx += 1
    data[:, col_idx] = raw_tap_density; col_idx += 1

    for i in range(36):
        data[:, col_idx] = np.random.normal(200 + i * 10, 20, ROWS)
        col_idx += 1  # 到 col 300

    assert col_idx == COLS, f"Expected {COLS} columns, got {col_idx}"
    return data


def generate_column_names():
    """生成 300 列的列名，按主题分组。"""
    names = []

    # 基础索引 (1-5)
    names += ["样本编号", "批次号", "设备编号", "班次", "基准环境温度"]

    # 涂布工序 (6-45)
    names += [
        "浆料粘度", "固含量", "涂布速度", "涂布间隙", "烘干温度",
        "烘干速度", "张力", "涂布厚度", "涂布面密度", "涂层密度",
        "表面粗糙度", "烘干速率", "残余水分", "孔隙率", "附着力",
    ]
    names += [f"涂布参数_{chr(65+i)}" for i in range(25)]

    # 辊压工序 (46-85)
    names += [
        "辊压压力", "辊压速度", "辊压温度", "辊压间隙", "辊压角度",
        "压实密度", "压实厚度", "延伸率", "表面光洁度", "硬度",
        "电导率",
    ]
    names += [f"辊压参数_{chr(65+i)}" for i in range(29)]

    # 装配工序 (86-125)
    names += [
        "焊接电流", "焊接电压", "焊接时间", "焊接压力", "极耳宽度",
        "超声波功率", "焊接电阻", "焊接强度", "密封宽度", "密封深度",
        "连接扭矩",
    ]
    names += [f"装配参数_{chr(65+i)}" for i in range(29)]

    # 注液工序 (126-155)
    names += [
        "注液量", "注液速度", "注液压力", "注液温度", "真空度",
        "保压时间", "浸润时间", "吸收率", "保液率",
    ]
    names += [f"注液参数_{chr(65+i)}" for i in range(21)]

    # 化成工序 (156-185)
    names += [
        "化成电流", "化成电压", "化成温度", "化成时间", "截止电压",
        "SEI阻抗", "初始容量", "库仑效率",
    ]
    names += [f"化成参数_{chr(65+i)}" for i in range(22)]

    # 品质检测 (186-225)
    names += [
        "开路电压", "内阻", "额定容量", "自放电率", "100周容量保持",
        "500周容量保持", "厚度变化", "重量",
    ]
    names += [f"品质指标_{chr(65+i)}" for i in range(32)]

    # 环境监测 (226-255)
    names += [
        "环境湿度", "环境温度", "粉尘浓度", "大气压", "振动幅度",
        "体感温度", "露点温度",
    ]
    names += [f"环境监测_{chr(65+i)}" for i in range(23)]

    # 原材料 (256-300)
    names += [
        "原料粒径", "BET比表面积", "原料纯度", "pH值", "原料水分",
        "原料电导率", "分散性", "流动性", "振实密度",
    ]
    names += [f"原料参数_{chr(65+i)}" for i in range(36)]

    assert len(names) == COLS, f"Expected {COLS} names, got {len(names)}"
    return names


def inject_missing_values(data, missing_rate=0.03):
    """按 3% 概率注入缺失值（用 NaN 表示）。"""
    mask = np.random.random(data.shape) < missing_rate
    data_float = data.astype(float)
    data_float[mask] = np.nan
    return data_float


def write_excel(filepath, data, col_names):
    """写入 Excel 文件，带表头样式。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "演示数据"

    # 写入表头
    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j, name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=j, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 写入数据
    for i in range(ROWS):
        for j in range(COLS):
            val = data[i, j]
            if np.isnan(val):
                ws.cell(row=i + 2, column=j + 1, value=None)
            else:
                ws.cell(row=i + 2, column=j + 1, value=round(float(val), 4))

    # 设置列宽
    for j in range(1, COLS + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14

    # 冻结首行
    ws.freeze_panes = "A2"

    # 设置数值格式（保留4位小数）
    for row in ws.iter_rows(min_row=2, max_row=ROWS + 1, min_col=1, max_col=COLS):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "0.0000"

    wb.save(filepath)
    print(f"已生成: {filepath}")
    print(f"  行数: {ROWS}, 列数: {COLS}")
    print(f"  包含约 3% 缺失值")


if __name__ == "__main__":
    output = "D:/FrontProjects/correlation-analysis-system/test/resource/demo_data_2000x300.xlsx"
    col_names = generate_column_names()
    raw_data = generate_data()
    data = inject_missing_values(raw_data, missing_rate=0.03)
    write_excel(output, data, col_names)
